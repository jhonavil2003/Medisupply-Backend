"""
Export service for sales reports.
Handles PDF and Excel export functionality.
"""
import os
import tempfile
from datetime import datetime
from typing import Dict, List, Any
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors


class ReportExportService:
    """Service for exporting sales reports to PDF and Excel."""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def export_to_excel(self, report_data: Dict[str, Any], filters: Dict[str, Any] = None) -> str:
        """
        Export report data to Excel file.
        
        Args:
            report_data: Dictionary containing summary, totals, etc.
            filters: Applied filters information
            
        Returns:
            str: Path to the created Excel file
        """
        # Prepare data
        df = pd.DataFrame(report_data.get('summary', []))
        
        if df.empty:
            # Create empty dataframe with headers
            df = pd.DataFrame(columns=[
                'Fecha', 'Vendedor', 'Región', 'Territorio', 'SKU Producto', 
                'Nombre Producto', 'Volumen Ventas', 'Valor Total', 
                'Objetivo', 'Tipo Objetivo', 'Cumplimiento %'
            ])
        else:
            # Rename columns to Spanish
            column_mapping = {
                'fecha': 'Fecha',
                'vendedor': 'Vendedor',
                'region': 'Región',
                'territory': 'Territorio',
                'product_sku': 'SKU Producto',
                'product_name': 'Nombre Producto',
                'volumen_ventas': 'Volumen Ventas',
                'valor_total': 'Valor Total',
                'valor_objetivo': 'Objetivo',
                'tipo_objetivo': 'Tipo Objetivo',
                'cumplimiento_porcentaje': 'Cumplimiento %'
            }
            
            # Select and rename columns
            available_cols = [col for col in column_mapping.keys() if col in df.columns]
            df = df[available_cols].rename(columns=column_mapping)
            
            # Format numeric columns
            if 'Valor Total' in df.columns:
                df['Valor Total'] = df['Valor Total'].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "")
            if 'Objetivo' in df.columns:
                df['Objetivo'] = df['Objetivo'].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
            if 'Cumplimiento %' in df.columns:
                df['Cumplimiento %'] = df['Cumplimiento %'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")
        
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reporte_ventas_{timestamp}.xlsx'
        filepath = os.path.join(self.temp_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Reporte de Ventas', index=False, startrow=4)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Reporte de Ventas']
            
            # Add header information
            self._add_excel_header(worksheet, filters, report_data.get('totals', {}))
            
            # Format the table
            self._format_excel_table(worksheet, df, start_row=5)
        
        return filepath
    
    def _add_excel_header(self, worksheet, filters: Dict, totals: Dict):
        """Add header information to Excel worksheet."""
        # Title
        worksheet['A1'] = 'REPORTE DE VENTAS - MEDISUPPLY'
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Date
        worksheet['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(size=10, italic=True)
        
        # Filters applied
        if filters:
            filter_text = "Filtros aplicados: " + ", ".join([f"{k}: {v}" for k, v in filters.items()])
            worksheet['A3'] = filter_text
            worksheet['A3'].font = Font(size=10, color="666666")
        
        # Summary totals - show in separate area
        if totals:
            col = 'H'  # Start from column H
            worksheet[f'{col}1'] = 'RESUMEN'
            worksheet[f'{col}1'].font = Font(size=12, bold=True)
            
            row = 2
            for key, value in totals.items():
                label = self._format_total_label(key)
                if isinstance(value, (int, float)):
                    if 'valor' in key.lower():
                        formatted_value = f"${value:,.2f}"
                    else:
                        formatted_value = f"{value:,}"
                else:
                    formatted_value = str(value)
                
                worksheet[f'{col}{row}'] = f"{label}: {formatted_value}"
                worksheet[f'{col}{row}'].font = Font(size=10)
                row += 1
    
    def _format_total_label(self, key: str) -> str:
        """Format total labels in Spanish."""
        labels = {
            'total_volumen_ventas': 'Total Unidades',
            'total_valor_total': 'Total Ventas',
            'unique_salespersons': 'Vendedores',
            'unique_products': 'Productos',
            'unique_regions': 'Regiones'
        }
        return labels.get(key, key.replace('_', ' ').title())
    
    def _format_excel_table(self, worksheet, df, start_row: int):
        """Apply formatting to the Excel table."""
        if df.empty:
            return
            
        # Header row formatting
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows - alternate colors
        light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        for row_num in range(start_row + 1, start_row + len(df) + 1):
            if (row_num - start_row) % 2 == 0:
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width
    
    def export_to_pdf(self, report_data: Dict[str, Any], filters: Dict[str, Any] = None) -> str:
        """
        Export report data to PDF file.
        
        Args:
            report_data: Dictionary containing summary, totals, etc.
            filters: Applied filters information
            
        Returns:
            str: Path to the created PDF file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reporte_ventas_{timestamp}.pdf'
        filepath = os.path.join(self.temp_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=0.5*inch)
        
        # Container for elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=HexColor('#2E7D32'),
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph("REPORTE DE VENTAS - MEDISUPPLY", title_style)
        elements.append(title)
        
        # Date and filters info
        date_text = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        if filters:
            filter_text = "Filtros aplicados: " + ", ".join([f"<b>{k}:</b> {v}" for k, v in filters.items()])
            elements.append(Paragraph(filter_text, styles['Normal']))
            elements.append(Spacer(1, 12))
        
        # Summary totals
        totals = report_data.get('totals', {})
        if totals:
            elements.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", styles['Heading2']))
            
            summary_data = []
            for key, value in totals.items():
                label = self._format_total_label(key)
                if isinstance(value, (int, float)):
                    if 'valor' in key.lower():
                        formatted_value = f"${value:,.2f}"
                    else:
                        formatted_value = f"{value:,}"
                else:
                    formatted_value = str(value)
                summary_data.append([label, formatted_value])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
        
        # Main data table
        summary_data = report_data.get('summary', [])
        if summary_data:
            elements.append(Paragraph("<b>DETALLE DE VENTAS</b>", styles['Heading2']))
            
            # Prepare table data
            headers = ['Fecha', 'Vendedor', 'Región', 'Producto', 'Cantidad', 'Valor', 'Objetivo', 'Cumpl.%']
            table_data = [headers]
            
            for item in summary_data:
                row = [
                    item.get('fecha', '')[:10] if item.get('fecha') else '',
                    item.get('vendedor', '')[:20] if item.get('vendedor') else '',
                    item.get('region', ''),
                    item.get('product_sku', ''),
                    str(item.get('volumen_ventas', 0)),
                    f"${item.get('valor_total', 0):,.0f}",
                    str(item.get('valor_objetivo', '') or ''),
                    f"{item.get('cumplimiento_porcentaje', 0):.1f}%" if item.get('cumplimiento_porcentaje') else ''
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
        else:
            elements.append(Paragraph("No se encontraron datos para los filtros aplicados.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    
    def cleanup_file(self, filepath: str) -> bool:
        """
        Clean up temporary file.
        
        Args:
            filepath: Path to file to delete
            
        Returns:
            bool: True if file was deleted successfully
        """
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
                return True
            return False
        except Exception:
            return False